"""Extraccion de registros de convenio desde las matrices anuales (solo lectura).

Para cada hoja de cada matriz: detecta la fila de encabezado, mapea columnas
reconocidas a un esquema comun (mapeo_columnas.py), descarta filas que no son
datos (pies de pagina tipo "Año:", "Elaborado por:"), y produce un dict por
fila de convenio con trazabilidad completa hasta el archivo/hoja/fila original.

NO modifica los archivos de matriz. Solo los abre con openpyxl en modo lectura.
"""

import re
from datetime import datetime
from pathlib import Path

import openpyxl

from mapeo_columnas import clasificar_tipo, es_fila_no_dato, mapear_encabezados

FILAS_MAX_BUSQUEDA_ENCABEZADO = 20
RE_FECHA_TEXTO = re.compile(r"(\d{1,2})[/\\-](\d{1,2})[/\\-](\d{2,4})")
RE_DIGITOS_FINALES = re.compile(r"(\d+)\D*$")


def _fila_parece_encabezado(valores: list) -> bool:
    no_vacias = [v for v in valores if v not in (None, "")]
    textuales = [v for v in no_vacias if isinstance(v, str)]
    return len(no_vacias) >= 3 and len(textuales) >= max(2, len(no_vacias) // 2)


def _parsear_fecha(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if hasattr(valor, "year") and hasattr(valor, "month"):  # date
        return valor
    texto = str(valor).strip()
    m = RE_FECHA_TEXTO.search(texto)
    if m:
        d, mo, y = m.groups()
        y = int(y)
        if y < 100:
            y += 2000
        try:
            return datetime(y, int(mo), int(d)).date()
        except ValueError:
            return None
    return None


def _texto(valor) -> str:
    if valor is None:
        return ""
    return str(valor).strip()


def _extraer_numero_original(codigo: str):
    if not codigo:
        return None
    m = RE_DIGITOS_FINALES.search(codigo)
    return m.group(1) if m else None


def _detectar_fila_encabezado(ws):
    for idx, fila in enumerate(ws.iter_rows(min_row=1, max_row=FILAS_MAX_BUSQUEDA_ENCABEZADO, values_only=True), start=1):
        if _fila_parece_encabezado(list(fila)):
            return idx, list(fila)
    return None, None


def extraer_registros_matriz(ruta: Path, anio: int, nombre_archivo: str):
    """Generador de dicts, uno por fila de convenio encontrada en cualquier hoja."""
    if ruta.suffix.lower() == ".xls":
        return  # formato antiguo no soportado por openpyxl, no aplica en este repositorio

    wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    try:
        for nombre_hoja in wb.sheetnames:
            ws = wb[nombre_hoja]
            fila_encabezado, encabezados_raw = _detectar_fila_encabezado(ws)
            if fila_encabezado is None:
                continue  # hoja sin encabezado reconocible (p.ej. hoja vacia) -> no se puede extraer, se reporta aparte

            mapa_bucket, mapa_original = mapear_encabezados(encabezados_raw)
            if not mapa_bucket:
                continue  # ninguna columna reconocida -> hoja probablemente vacia/formato distinto

            for idx_fila, fila in enumerate(ws.iter_rows(min_row=fila_encabezado + 1, values_only=True), start=fila_encabezado + 1):
                valores = list(fila)
                if es_fila_no_dato(valores):
                    continue

                datos = {}
                for idx_col, bucket in mapa_bucket.items():
                    if idx_col < len(valores):
                        datos[bucket] = valores[idx_col]

                # columnas no reconocidas: conservarlas para no perder informacion
                extras = []
                buckets_usados_idx = set(mapa_bucket.keys())
                for idx_col, texto_original_col in mapa_original.items():
                    if idx_col in buckets_usados_idx:
                        continue
                    if idx_col < len(valores) and valores[idx_col] not in (None, ""):
                        extras.append(f"{texto_original_col}: {valores[idx_col]}")

                codigo_original = _texto(datos.get("CODIGO")) or None
                institucion = _texto(datos.get("NOMBRE_COMERCIAL")) or _texto(datos.get("COMPARECIENTES")) or None
                subtipo_original = _texto(datos.get("NATURALEZA")) or None
                tipo_normalizado, clasificacion_general = clasificar_tipo(subtipo_original, nombre_hoja)

                observaciones = _texto(datos.get("OBSERVACIONES"))
                if extras:
                    sufijo = " | ".join(extras)
                    observaciones = f"{observaciones} | {sufijo}" if observaciones else sufijo

                registro = {
                    "anio": anio,
                    "codigo_original": codigo_original,
                    "numero_original": _extraer_numero_original(codigo_original),
                    "institucion": institucion,
                    "tipo_instrumento": tipo_normalizado,
                    "subtipo": subtipo_original,
                    "clasificacion_general": clasificacion_general,
                    "objeto": _texto(datos.get("OBJETO")) or None,
                    "fecha_suscripcion": _parsear_fecha(datos.get("FECHA_SUSCRIPCION")),
                    "plazo": _texto(datos.get("PERIODO_VIGENCIA")) or None,
                    "administrador": _texto(datos.get("ADMINISTRADOR")) or None,
                    "unidad_responsable": (
                        _texto(datos.get("UNIDAD_GESTORA"))
                        or _texto(datos.get("TRAMITE_PRESENTADO_POR"))
                        or None
                    ),
                    "observaciones_originales": observaciones or None,
                    "hoja_origen": nombre_hoja,
                    "archivo_matriz_origen": nombre_archivo,
                    "fila_origen": idx_fila,
                    # campos adicionales de valor (superconjunto del minimo pedido)
                    "ruc": _texto(datos.get("RUC")) or None,
                    "ambito": _texto(datos.get("AMBITO")) or None,
                    "seccion": _texto(datos.get("SECCION")) or None,
                    "sector": _texto(datos.get("SECTOR")) or None,
                    "direccion": _texto(datos.get("DIRECCION")) or None,
                    "representante_legal": _texto(datos.get("REPRESENTANTE_LEGAL")) or None,
                    "contacto": _texto(datos.get("CONTACTO")) or None,
                    "email": _texto(datos.get("EMAIL")) or None,
                    "telefono": _texto(datos.get("TELEFONO")) or None,
                    "carreras_beneficiadas": _texto(datos.get("CARRERAS_BENEFICIADAS")) or None,
                    "estado_original": _texto(datos.get("ESTADO_ORIGINAL")) or None,
                    "link_documento_matriz": _texto(datos.get("LINK_DOCUMENTO")) or None,
                }

                # Filtrar filas totalmente vacias de contenido util (sin codigo NI institucion NI objeto)
                if not (registro["codigo_original"] or registro["institucion"] or registro["objeto"]):
                    continue

                yield registro
    finally:
        wb.close()
