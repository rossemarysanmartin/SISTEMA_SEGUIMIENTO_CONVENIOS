"""Pruebas del modulo de Solicitudes y Trazabilidad (Fase 5).

IMPORTANTE: se ejecutan contra una base SQLite TEMPORAL (tempfile), nunca
contra BASE_DATOS/convenios.db real -- no se debe contaminar la base
institucional con datos de prueba (seccion 39 de la especificacion).
"""

import sqlite3
import sys
import tempfile
import threading
from datetime import date, timedelta
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import db_fase5
import db_fase6
import db_maestra
from services import repositorio_solicitudes as repo
from services.fechas_habiles import contar_dias_habiles


@pytest.fixture()
def conn():
    """Conexion a una base SQLite temporal, con el esquema de convenios
    (Fase 2) + solicitudes (Fase 5) ya migrado."""
    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp.close()
    ruta = Path(tmp.name)
    conexion = sqlite3.connect(str(ruta), isolation_level=None)
    conexion.row_factory = sqlite3.Row
    conexion.execute("PRAGMA foreign_keys = ON;")
    db_maestra.inicializar_esquema(conexion)
    db_fase5.migrar(conexion)
    db_fase6.migrar(conexion)
    yield conexion
    conexion.close()
    ruta.unlink(missing_ok=True)


def _datos_solicitud(**overrides):
    datos = {
        "fecha_ingreso": "2026-08-20",
        "hora_ingreso": "09:15",
        "institucion": "Gobierno Autónomo Descentralizado de Prueba",
        "persona_contacto": "Juan Pérez",
        "correo_contacto": "juan@example.com",
        "asunto": "Solicitud de convenio de cooperación",
        "tipo_convenio_solicitado": "COOPERACION",
        "dependencia_solicitante": "Rectorado",
        "medio_ingreso": "CORREO_ELECTRONICO",
        "observaciones": "Ninguna",
    }
    datos.update(overrides)
    return datos


# --------------------------------------------------------- codigo correlativo --

def test_codigo_correlativo_formato_y_secuencia(conn):
    c1 = repo.generar_codigo_solicitud(conn, 2026)
    c2 = repo.generar_codigo_solicitud(conn, 2026)
    assert c1 == "SOL-2026-0001"
    assert c2 == "SOL-2026-0002"


def test_codigo_correlativo_es_independiente_por_anio(conn):
    c2026 = repo.generar_codigo_solicitud(conn, 2026)
    c2027 = repo.generar_codigo_solicitud(conn, 2027)
    assert c2026 == "SOL-2026-0001"
    assert c2027 == "SOL-2027-0001"


def test_codigo_correlativo_no_se_duplica_bajo_concurrencia(conn):
    """Simula varios hilos generando codigos al mismo tiempo (misma conexion
    no es realista para SQLite multihilo, asi que cada hilo abre su propia
    conexion a la MISMA base temporal, que es lo que importaria en un
    escenario real)."""
    ruta = Path(conn.execute("PRAGMA database_list").fetchone()[2])
    codigos = []
    lock = threading.Lock()
    errores = []

    def trabajo():
        try:
            c = sqlite3.connect(str(ruta), isolation_level=None, timeout=5)
            c.row_factory = sqlite3.Row
            codigo = repo.generar_codigo_solicitud(c, 2026)
            with lock:
                codigos.append(codigo)
            c.close()
        except Exception as exc:
            errores.append(exc)

    n = 5
    hilos = [threading.Thread(target=trabajo) for _ in range(n)]
    for h in hilos:
        h.start()
    for h in hilos:
        h.join()

    assert not errores
    assert len(codigos) == len(set(codigos)) == n


def test_crear_solicitud_no_permite_codigo_duplicado_manual(conn):
    creada = repo.crear_solicitud(conn, _datos_solicitud())
    with pytest.raises(sqlite3.IntegrityError):
        conn.execute("BEGIN IMMEDIATE")
        conn.execute(
            "INSERT INTO solicitudes (codigo_solicitud, anio, fecha_ingreso, institucion, medio_ingreso, "
            "estado_actual, activo, fecha_creacion, fecha_actualizacion) VALUES (?,2026,'2026-01-01','X','OTRO','RECIBIDA',1,'x','x')",
            (creada["codigo_solicitud"],),
        )
        conn.execute("COMMIT")


# ------------------------------------------------------------- solicitudes --

def test_crear_solicitud_valores_iniciales(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    assert s["codigo_solicitud"] == "SOL-2026-0001"
    assert s["estado_actual"] == "RECIBIDA"
    assert s["etapa_actual"] == "RECEPCION"
    assert s["activo"] == 1


def test_crear_solicitud_con_recepcion_en_vinculacion_genera_primera_actuacion(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud(registrado_por="María López"), recibido_en_vinculacion=True)
    actuaciones = repo.obtener_actuaciones(conn, s["id"])
    assert len(actuaciones) == 1
    assert actuaciones[0]["tipo_actuacion"] == "RECEPCION"
    assert actuaciones[0]["dependencia_destino"] == "Dirección de Vinculación"


def test_registrar_traslado_actualiza_etapa_y_solicitud(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    repo.registrar_actuacion(conn, s["id"], {
        "tipo_actuacion": "TRASLADO",
        "dependencia_origen": "Dirección de Vinculación",
        "dependencia_destino": "Unidad de Relaciones Interinstitucionales",
        "etapa_nueva": "RECEPCION",
        "estado_nuevo": "EN_GESTION",
    })
    actualizada = repo.obtener_solicitud(conn, s["id"])
    assert actualizada["estado_actual"] == "EN_GESTION"
    assert actualizada["fecha_ultima_actuacion"] is not None


def test_delegacion_actualiza_delegado_actual_y_conserva_historial(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    repo.registrar_actuacion(conn, s["id"], {"tipo_actuacion": "DELEGACION", "delegado": "Ana Torres"})
    repo.registrar_actuacion(conn, s["id"], {"tipo_actuacion": "DELEGACION", "delegado": "Carlos Ruiz"})
    actualizada = repo.obtener_solicitud(conn, s["id"])
    assert actualizada["delegado_actual"] == "Carlos Ruiz"
    historial = repo.obtener_actuaciones(conn, s["id"])
    delegados_historicos = [a["delegado"] for a in historial if a["tipo_actuacion"] == "DELEGACION"]
    assert delegados_historicos == ["Ana Torres", "Carlos Ruiz"]  # el historial no se borra


def test_solicitud_de_criterio_queda_pendiente_de_respuesta(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    fecha_envio = (date.today() - timedelta(days=10)).isoformat()
    repo.registrar_actuacion(conn, s["id"], {
        "tipo_actuacion": "SOLICITUD_DE_CRITERIO",
        "fecha": fecha_envio,
        "dependencia_destino": "Procuraduría / Asesoría Jurídica",
        "estado_nuevo": "PENDIENTE_DE_CRITERIO",
        "requiere_respuesta": "SI",
    })
    pendientes = repo.pendientes_de_respuesta(conn, umbral_dias_habiles=5)
    assert len(pendientes) == 1
    assert pendientes[0]["id_solicitud"] == s["id"]
    assert pendientes[0]["dias_habiles_esperando"] > 0
    assert pendientes[0]["alerta"] is True


def test_respuesta_de_criterio_marca_actuacion_original_como_respondida(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    id_solicitud_criterio = repo.registrar_actuacion(conn, s["id"], {
        "tipo_actuacion": "SOLICITUD_DE_CRITERIO",
        "dependencia_destino": "Procuraduría / Asesoría Jurídica",
        "requiere_respuesta": "SI",
    })
    repo.registrar_actuacion(conn, s["id"], {
        "tipo_actuacion": "CRITERIO_RECIBIDO",
        "resultado": "FAVORABLE",
        "id_actuacion_relacionada": id_solicitud_criterio,
        "respuesta_recibida": "SI",
    })
    pendientes = repo.pendientes_de_respuesta(conn, umbral_dias_habiles=5)
    assert len(pendientes) == 0
    original = conn.execute("SELECT * FROM actuaciones_solicitud WHERE id=?", (id_solicitud_criterio,)).fetchone()
    assert original["respuesta_recibida"] == "SI"


def test_dias_sin_movimiento_se_reinicia_con_nueva_actuacion(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud(fecha_ingreso=(date.today() - timedelta(days=20)).isoformat()))
    repo.registrar_actuacion(conn, s["id"], {"tipo_actuacion": "REVISION_INICIAL", "fecha": date.today().isoformat()})
    actualizada = repo.obtener_solicitud(conn, s["id"])
    assert actualizada["dias_sin_movimiento"] == 0


def test_semaforo_umbrales_configurables():
    class ConfigFalsa:
        semaforo_normal_max = 3
        semaforo_atencion_max = 7
        semaforo_demora_max = 14
    cfg = ConfigFalsa()
    assert repo.calcular_semaforo(0, cfg)["codigo"] == "NORMAL"
    assert repo.calcular_semaforo(5, cfg)["codigo"] == "ATENCION"
    assert repo.calcular_semaforo(10, cfg)["codigo"] == "DEMORA"
    assert repo.calcular_semaforo(20, cfg)["codigo"] == "REVISAR"


def test_dias_habiles_excluye_fin_de_semana():
    # Un viernes a un lunes: 1 dia habil (el lunes), no 3 dias calendario
    viernes = date(2026, 8, 21)
    lunes = date(2026, 8, 24)
    assert contar_dias_habiles(viernes, lunes) == 1


# ---------------------------------------------------------------- filtros --

def test_filtro_por_estado(conn):
    repo.crear_solicitud(conn, _datos_solicitud(institucion="Institución A"))
    s2 = repo.crear_solicitud(conn, _datos_solicitud(institucion="Institución B"))
    repo.registrar_actuacion(conn, s2["id"], {"tipo_actuacion": "REVISION_INICIAL", "estado_nuevo": "EN_GESTION"})

    filas, total, _ = repo.listar_solicitudes(conn, {"estado_actual": "EN_GESTION"}, "", "fecha_ingreso", "desc", 1)
    assert total == 1
    assert filas[0]["institucion"] == "Institución B"


def test_busqueda_case_insensitive(conn):
    repo.crear_solicitud(conn, _datos_solicitud(institucion="Universidad Católica"))
    filas_min, total_min, _ = repo.listar_solicitudes(conn, {}, "universidad", "fecha_ingreso", "desc", 1)
    filas_mayus, total_mayus, _ = repo.listar_solicitudes(conn, {}, "UNIVERSIDAD", "fecha_ingreso", "desc", 1)
    assert total_min == total_mayus == 1


# --------------------------------------------------------------- auditoria --

def test_cambio_de_estado_genera_auditoria(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    repo.registrar_actuacion(conn, s["id"], {"tipo_actuacion": "REVISION_INICIAL", "estado_nuevo": "EN_GESTION"})
    entradas = conn.execute(
        "SELECT * FROM auditoria WHERE entidad='solicitud' AND id_entidad=? AND accion='CAMBIO_ESTADO'", (s["id"],)
    ).fetchall()
    assert len(entradas) == 1
    assert entradas[0]["valor_anterior"] == "RECIBIDA"
    assert entradas[0]["valor_nuevo"] == "EN_GESTION"


def test_delegacion_genera_auditoria(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    repo.registrar_actuacion(conn, s["id"], {"tipo_actuacion": "DELEGACION", "delegado": "Ana Torres"})
    entradas = conn.execute(
        "SELECT * FROM auditoria WHERE entidad='solicitud' AND accion='DELEGACION' AND id_entidad=?", (s["id"],)
    ).fetchall()
    assert len(entradas) == 1


# --------------------------------------------------------- vinculo convenio --

def test_vincular_convenio_es_bidireccional(conn):
    conn.execute(
        "INSERT INTO convenios (anio, institucion, clasificacion_general) VALUES (2026, 'Institución X', 'CONVENIO')"
    )
    id_convenio = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    s = repo.crear_solicitud(conn, _datos_solicitud())

    repo.vincular_convenio(conn, s["id"], id_convenio)

    solicitud = repo.obtener_solicitud(conn, s["id"])
    convenio = conn.execute("SELECT * FROM convenios WHERE id_sistema=?", (id_convenio,)).fetchone()
    assert solicitud["id_convenio_suscrito"] == id_convenio
    assert convenio["id_solicitud_origen"] == s["id"]


# -------------------------------------------------------- eliminacion logica --

def test_desactivar_solicitud_no_borra_fisicamente(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    repo.desactivar_solicitud(conn, s["id"], motivo="Duplicada")
    fila = conn.execute("SELECT * FROM solicitudes WHERE id=?", (s["id"],)).fetchone()
    assert fila is not None  # sigue existiendo
    assert fila["activo"] == 0
    assert fila["estado_actual"] == "ARCHIVADO"


def test_correccion_no_modifica_actuacion_original_en_silencio(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    id_original = repo.registrar_actuacion(conn, s["id"], {"tipo_actuacion": "REVISION_INICIAL", "descripcion": "texto original"})
    repo.registrar_correccion(conn, s["id"], id_original, motivo="Error de tipeo en la descripción")

    original = conn.execute("SELECT * FROM actuaciones_solicitud WHERE id=?", (id_original,)).fetchone()
    assert original["descripcion"] == "texto original"  # no se toco

    correccion = conn.execute(
        "SELECT * FROM actuaciones_solicitud WHERE tipo_actuacion='CORRECCION' AND id_actuacion_relacionada=?",
        (id_original,),
    ).fetchone()
    assert correccion is not None


# --------------------------------------------------------- ruta relativa docs --

def test_documento_tramite_guarda_ruta_y_metadatos(conn):
    s = repo.crear_solicitud(conn, _datos_solicitud())
    repo.registrar_documento_tramite(conn, s["id"], {
        "nombre": "oficio_recibido.pdf",
        "ruta": r"SISTEMA_SEGUIMIENTO_CONVENIOS\TRAMITES\2026\SOL-2026-0001\oficio_recibido.pdf",
        "tipo": "OFICIO",
        "fecha": "2026-08-20",
    })
    documentos = repo.obtener_documentos_tramite(conn, s["id"])
    assert len(documentos) == 1
    assert documentos[0]["nombre"] == "oficio_recibido.pdf"
    assert "TRAMITES" in documentos[0]["ruta"]


def test_documento_tramite_calcula_ruta_relativa_al_repositorio_raiz(conn, tmp_path):
    """Seccion 35: ademas de la ruta absoluta, se conserva la ruta relativa al
    repositorio raiz, para no depender de esa ruta absoluta local especifica."""
    raiz = tmp_path / "CONVENIOS INTERINSTITUCIONALES UTMACH"
    absoluta = raiz / "SISTEMA_SEGUIMIENTO_CONVENIOS" / "TRAMITES" / "2026" / "SOL-2026-0001" / "oficio.pdf"
    s = repo.crear_solicitud(conn, _datos_solicitud())
    repo.registrar_documento_tramite(
        conn, s["id"],
        {"nombre": "oficio.pdf", "ruta": str(absoluta), "tipo": "OFICIO", "fecha": "2026-08-20"},
        ruta_base_convenios=raiz,
    )
    documento = repo.obtener_documentos_tramite(conn, s["id"])[0]
    assert documento["ruta_relativa"] == str(Path("SISTEMA_SEGUIMIENTO_CONVENIOS") / "TRAMITES" / "2026" / "SOL-2026-0001" / "oficio.pdf")


# --------------------------------------------------------------- exportacion --

def test_exportacion_excel_genera_las_cuatro_hojas(conn, tmp_path):
    from services.exportar_excel_solicitudes import generar_excel_solicitudes

    s = repo.crear_solicitud(conn, _datos_solicitud())
    repo.registrar_actuacion(conn, s["id"], {"tipo_actuacion": "REVISION_INICIAL", "estado_nuevo": "EN_GESTION"})

    ruta_db = Path(conn.execute("PRAGMA database_list").fetchone()[2])
    ruta_salida = tmp_path / "REPORTE_SOLICITUDES.xlsx"
    generar_excel_solicitudes(ruta_db, ruta_salida)

    assert ruta_salida.exists()
    from openpyxl import load_workbook
    wb = load_workbook(str(ruta_salida))
    assert set(wb.sheetnames) == {"SOLICITUDES", "TRAZABILIDAD", "PENDIENTES", "RESUMEN"}
    assert wb["SOLICITUDES"].max_row == 2  # encabezado + 1 solicitud
    assert wb["TRAZABILIDAD"].max_row == 2  # encabezado + 1 actuacion
