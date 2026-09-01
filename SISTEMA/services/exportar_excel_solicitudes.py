"""Genera BASE_DATOS/REPORTE_SOLICITUDES.xlsx a partir de convenios.db.

Artefacto GENERADO (no un documento original); se puede regenerar en
cualquier momento desde el boton "Exportar solicitudes". No sobrescribe
ninguna matriz historica ni el Excel maestro de convenios.
"""

import sqlite3
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from exportar_excel import _escribir_hoja
from services import repositorio_solicitudes as repo
from services.fechas_habiles import contar_dias_habiles


def generar_excel_solicitudes(ruta_db: Path, ruta_salida: Path, umbral_dias_habiles: int = 5):
    conn = sqlite3.connect(str(ruta_db))
    conn.row_factory = sqlite3.Row

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "SOLICITUDES"
    encabezados1 = [
        "Código", "Año", "Fecha ingreso", "Institución", "Medio ingreso", "Tipo solicitado",
        "Dependencia solicitante", "Responsable actual", "Delegado actual", "Etapa actual",
        "Estado actual", "Fecha última actuación", "Días sin movimiento", "Convenio vinculado", "Activo",
    ]
    filas1 = []
    for s in conn.execute("SELECT * FROM solicitudes ORDER BY anio DESC, codigo_solicitud DESC"):
        filas1.append([
            s["codigo_solicitud"], s["anio"], s["fecha_ingreso"], s["institucion"], s["medio_ingreso"],
            s["tipo_convenio_solicitado"], s["dependencia_solicitante"], s["responsable_actual"],
            s["delegado_actual"], s["etapa_actual"], s["estado_actual"], s["fecha_ultima_actuacion"],
            repo.dias_desde(s["fecha_ultima_actuacion"]), s["id_convenio_suscrito"], "SI" if s["activo"] else "NO",
        ])
    _escribir_hoja(ws1, encabezados1, filas1)

    ws2 = wb.create_sheet("TRAZABILIDAD")
    encabezados2 = [
        "Código solicitud", "Fecha", "Hora", "Actuación", "Dependencia origen", "Dependencia destino",
        "Responsable", "Delegado", "Estado anterior", "Estado nuevo", "Etapa anterior", "Etapa nueva",
        "Requiere respuesta", "Respuesta recibida", "Descripción",
    ]
    filas2 = []
    for a in conn.execute(
        """SELECT s.codigo_solicitud, a.* FROM actuaciones_solicitud a
           JOIN solicitudes s ON s.id = a.id_solicitud
           ORDER BY s.codigo_solicitud, a.fecha, a.hora, a.id"""
    ):
        filas2.append([
            a["codigo_solicitud"], a["fecha"], a["hora"], a["tipo_actuacion"], a["dependencia_origen"],
            a["dependencia_destino"], a["responsable"], a["delegado"], a["estado_anterior"], a["estado_nuevo"],
            a["etapa_anterior"], a["etapa_nueva"], a["requiere_respuesta"], a["respuesta_recibida"],
            (a["descripcion"] or "")[:300],
        ])
    _escribir_hoja(ws2, encabezados2, filas2)

    ws3 = wb.create_sheet("PENDIENTES")
    encabezados3 = ["Código solicitud", "Institución", "Actuación", "Dependencia destino", "Responsable",
                    "Fecha envío", "Fecha límite", "Días hábiles esperando", "Alerta"]
    filas3 = []
    for p in repo.pendientes_de_respuesta(conn, umbral_dias_habiles):
        filas3.append([
            p["codigo_solicitud"], p["institucion"], p["tipo_actuacion"], p["dependencia_destino"],
            p["responsable"], p["fecha_envio"], p["fecha_limite_respuesta"], p["dias_habiles_esperando"],
            "SI" if p["alerta"] else "NO",
        ])
    _escribir_hoja(ws3, encabezados3, filas3)

    ws4 = wb.create_sheet("RESUMEN")
    encabezados4 = ["Indicador", "Valor"]
    filas4 = [["Total solicitudes activas", conn.execute("SELECT COUNT(*) FROM solicitudes WHERE activo=1").fetchone()[0]]]
    for fila in repo.informe_por_estado(conn):
        filas4.append([f"Estado: {fila['estado_actual']}", fila["total"]])
    for fila in repo.informe_por_etapa(conn):
        filas4.append([f"Etapa: {fila['etapa_actual']}", fila["total"]])
    for fila in repo.informe_por_medio_ingreso(conn):
        filas4.append([f"Medio: {fila['medio_ingreso']}", fila["total"]])
    filas4.append(["Generado", date.today().isoformat()])
    _escribir_hoja(ws4, encabezados4, filas4)

    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_salida))
    conn.close()
