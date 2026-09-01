"""Inspeccion de solo lectura sobre las matrices anuales de convenios (.xlsx/.xls/.xlsm).

Nunca se abre un libro en modo escritura, nunca se guarda ningun cambio.
"""

import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    openpyxl = None

FILAS_MAX_BUSQUEDA_ENCABEZADO = 20


def _fila_parece_encabezado(valores: list) -> bool:
    no_vacias = [v for v in valores if v not in (None, "")]
    textuales = [v for v in no_vacias if isinstance(v, str)]
    return len(no_vacias) >= 3 and len(textuales) >= max(2, len(no_vacias) // 2)


def inspeccionar_matriz(ruta: Path) -> dict:
    """Devuelve {"hojas": [...], "detalle_hojas": [...], "notas": str} sin modificar el archivo."""
    if openpyxl is None:
        return {"hojas": [], "detalle_hojas": [], "notas": "openpyxl no esta instalado"}

    if ruta.suffix.lower() == ".xls":
        return {"hojas": [], "detalle_hojas": [], "notas": "Formato .xls antiguo no soportado por openpyxl (requiere xlrd)"}

    try:
        wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    except Exception as e:
        return {"hojas": [], "detalle_hojas": [], "notas": f"ERROR_LECTURA: {e}"}

    hojas = wb.sheetnames
    detalle_hojas = []

    for nombre_hoja in hojas:
        try:
            ws = wb[nombre_hoja]
            fila_encabezado = None
            encabezados = []
            filas_leidas = 0

            for idx, fila in enumerate(ws.iter_rows(min_row=1, max_row=FILAS_MAX_BUSQUEDA_ENCABEZADO, values_only=True), start=1):
                filas_leidas = idx
                if _fila_parece_encabezado(list(fila)):
                    fila_encabezado = idx
                    encabezados = [str(v).strip() if v is not None else "" for v in fila]
                    break

            filas_con_datos = 0
            if fila_encabezado is not None:
                for fila in ws.iter_rows(min_row=fila_encabezado + 1, values_only=True):
                    if any(v not in (None, "") for v in fila):
                        filas_con_datos += 1

            detalle_hojas.append({
                "nombre_hoja": nombre_hoja,
                "fila_encabezado_estimada": fila_encabezado,
                "encabezados_json": json.dumps(encabezados, ensure_ascii=False),
                "filas_con_datos_aprox": filas_con_datos,
            })
        except Exception as e:
            detalle_hojas.append({
                "nombre_hoja": nombre_hoja,
                "fila_encabezado_estimada": None,
                "encabezados_json": json.dumps([]),
                "filas_con_datos_aprox": 0,
                "error": str(e),
            })

    wb.close()
    return {"hojas": hojas, "detalle_hojas": detalle_hojas, "notas": None}
