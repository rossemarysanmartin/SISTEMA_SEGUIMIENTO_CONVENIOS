"""Inspeccion detallada de solo lectura de las matrices anuales.

Para cada matriz y cada hoja, imprime: fila de encabezado detectada, lista
completa de encabezados, y hasta 5 filas de datos de ejemplo. Esto se usa
UNA VEZ para disenar el mapeo de columnas (mapeo_columnas.py) con datos
reales, en vez de adivinar la estructura.

No modifica ningun archivo. Escribe su salida a REPORTES/INSPECCION_DETALLADA_MATRICES.md
"""

import json
from pathlib import Path

import openpyxl

from config import cargar_config

FILAS_MUESTRA = 5
FILAS_MAX_BUSQUEDA_ENCABEZADO = 20


def _fila_parece_encabezado(valores: list) -> bool:
    no_vacias = [v for v in valores if v not in (None, "")]
    textuales = [v for v in no_vacias if isinstance(v, str)]
    return len(no_vacias) >= 3 and len(textuales) >= max(2, len(no_vacias) // 2)


def inspeccionar(ruta: Path, lineas: list):
    lineas.append(f"\n## {ruta.name}\n")
    try:
        wb = openpyxl.load_workbook(str(ruta), read_only=True, data_only=True)
    except Exception as e:
        lineas.append(f"ERROR al abrir: {e}\n")
        return

    for nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        lineas.append(f"\n### Hoja: {nombre_hoja}\n")

        fila_encabezado = None
        encabezados = []
        filas_buffer = []
        for idx, fila in enumerate(ws.iter_rows(min_row=1, max_row=FILAS_MAX_BUSQUEDA_ENCABEZADO, values_only=True), start=1):
            filas_buffer.append((idx, fila))
            if fila_encabezado is None and _fila_parece_encabezado(list(fila)):
                fila_encabezado = idx
                encabezados = list(fila)

        lineas.append(f"- Fila de encabezado estimada: {fila_encabezado}\n")
        lineas.append(f"- Encabezados: {json.dumps([str(h) if h is not None else None for h in encabezados], ensure_ascii=False)}\n")

        lineas.append("\nFilas previas al encabezado (contexto):\n")
        for idx, fila in filas_buffer:
            if fila_encabezado and idx >= fila_encabezado:
                break
            no_vacias = [v for v in fila if v not in (None, "")]
            if no_vacias:
                lineas.append(f"  fila {idx}: {no_vacias}\n")

        lineas.append("\nMuestra de filas de datos:\n")
        if fila_encabezado is not None:
            contador = 0
            for fila in ws.iter_rows(min_row=fila_encabezado + 1, values_only=True):
                if contador >= FILAS_MUESTRA:
                    break
                if any(v not in (None, "") for v in fila):
                    lineas.append(f"  {list(fila)}\n")
                    contador += 1
        lineas.append("\n")

    wb.close()


def main():
    config = cargar_config()
    lineas = ["# Inspeccion detallada de matrices - Solo lectura\n"]

    conn_rutas = []
    for anio in config.anios_analizar:
        ruta_anio = config.ruta_base_convenios / str(anio)
        if not ruta_anio.is_dir():
            continue
        for archivo in sorted(ruta_anio.iterdir()):
            if archivo.is_file() and archivo.suffix.lower() in config.extensiones_matriz:
                conn_rutas.append(archivo)

    for ruta in conn_rutas:
        inspeccionar(ruta, lineas)

    salida = config.ruta_reportes / "INSPECCION_DETALLADA_MATRICES.md"
    salida.parent.mkdir(parents=True, exist_ok=True)
    with open(salida, "w", encoding="utf-8") as f:
        f.writelines(lineas)
    print(f"Inspeccion guardada en: {salida}")


if __name__ == "__main__":
    main()
