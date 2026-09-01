"""Genera BASE_MAESTRA_CONVENIOS_2020_2026.xlsx a partir de convenios.db.

Este Excel es un ARTEFACTO GENERADO por el sistema (no es un documento
original) y puede reemplazarse en cada sincronizacion futura.
"""

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

RELLENO_ENCABEZADO = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FUENTE_ENCABEZADO = Font(color="FFFFFF", bold=True)


def _escribir_hoja(ws, encabezados: list, filas: list):
    ws.append(encabezados)
    for col_idx in range(1, len(encabezados) + 1):
        celda = ws.cell(row=1, column=col_idx)
        celda.fill = RELLENO_ENCABEZADO
        celda.font = FUENTE_ENCABEZADO
    ws.freeze_panes = "A2"
    for fila in filas:
        ws.append(fila)
    if filas:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(encabezados))}{len(filas) + 1}"
    for col_idx, encabezado in enumerate(encabezados, start=1):
        max_len = len(str(encabezado))
        for fila in filas[:200]:  # muestreo para no recorrer todo en columnas de texto largo
            valor = fila[col_idx - 1]
            if valor is not None:
                max_len = max(max_len, min(len(str(valor)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 60))


def _columnas_disponibles(conn, tabla: str, deseadas: list) -> list:
    existentes = {fila[1] for fila in conn.execute(f"PRAGMA table_info({tabla})")}
    return [c for c in deseadas if c in existentes]


def generar_excel_maestro(ruta_db: Path, ruta_salida: Path):
    conn = sqlite3.connect(str(ruta_db))
    conn.row_factory = sqlite3.Row

    wb = Workbook()
    wb.remove(wb.active)

    # ---- CONVENIOS ----
    cols_convenios = _columnas_disponibles(conn, "convenios", [
        "id_sistema", "anio", "codigo_original", "numero_original", "institucion", "institucion_normalizada",
        "tipo_instrumento", "subtipo", "clasificacion_general", "objeto", "objeto_documento_resumen",
        "fecha_suscripcion", "fecha_suscripcion_documento", "fecha_inicio", "plazo", "plazo_documento",
        "fecha_finalizacion", "fecha_finalizacion_documento", "estado_vigencia", "dias_para_vencimiento",
        "estado_revision_vigencia", "administrador", "administrador_documento", "unidad_responsable",
        "estado_relacion_documental", "confianza_relacion", "ruta_documento_principal",
        "tiene_adenda", "documento_adenda_ruta", "conflicto_fecha", "requiere_revision",
        "requiere_revision_documental", "confianza_analisis", "fecha_ultimo_analisis_documento",
        "hoja_origen", "archivo_matriz_origen", "fila_origen",
    ])
    filas = [tuple(row[c] for c in cols_convenios) for row in conn.execute(f"SELECT {', '.join(cols_convenios)} FROM convenios ORDER BY anio, tipo_instrumento")]
    ws = wb.create_sheet("CONVENIOS")
    _escribir_hoja(ws, cols_convenios, filas)

    # ---- DOCUMENTOS ----
    cols_docs = _columnas_disponibles(conn, "documentos", [
        "id_documento", "id_convenio", "anio", "carpeta_tipo", "nombre", "extension", "tamano", "fecha_modificacion",
        "clasificacion_tecnica_pdf", "tipo_documento_tecnico", "firma_electronica_detectada", "cantidad_firmas",
        "firmante_certificado", "fecha_firma_metadato", "requiere_revision", "es_documento_principal", "ruta",
    ])
    filas = [tuple(row[c] for c in cols_docs) for row in conn.execute(f"SELECT {', '.join(cols_docs)} FROM documentos ORDER BY anio, carpeta_tipo")]
    ws = wb.create_sheet("DOCUMENTOS")
    _escribir_hoja(ws, cols_docs, filas)

    # ---- RESUMEN ----
    ws = wb.create_sheet("RESUMEN")
    resumen_filas = []
    total_convenios = conn.execute("SELECT COUNT(*) FROM convenios").fetchone()[0]
    resumen_filas.append(("Total de registros importados de matrices", total_convenios))
    for row in conn.execute("SELECT anio, COUNT(*) AS n FROM convenios GROUP BY anio ORDER BY anio"):
        resumen_filas.append((f"Registros año {row['anio']}", row["n"]))
    for row in conn.execute("SELECT clasificacion_general, COUNT(*) AS n FROM convenios GROUP BY clasificacion_general ORDER BY n DESC"):
        resumen_filas.append((f"Clasificación: {row['clasificacion_general']}", row["n"]))
    for row in conn.execute("SELECT estado_relacion_documental, COUNT(*) AS n FROM convenios GROUP BY estado_relacion_documental ORDER BY n DESC"):
        resumen_filas.append((f"Relación documental: {row['estado_relacion_documental']}", row["n"]))
    for row in conn.execute("SELECT estado_vigencia, COUNT(*) AS n FROM convenios GROUP BY estado_vigencia ORDER BY n DESC"):
        resumen_filas.append((f"Vigencia: {row['estado_vigencia']}", row["n"]))
    _escribir_hoja(ws, ["Indicador", "Valor"], resumen_filas)

    # ---- REVISION_MANUAL ----
    cols_rev = ["id_sistema", "anio", "codigo_original", "institucion", "tipo_instrumento", "clasificacion_general",
                "estado_relacion_documental", "notas_sistema"]
    filas = [tuple(row[c] for c in cols_rev) for row in conn.execute(
        f"SELECT {', '.join(cols_rev)} FROM convenios WHERE requiere_revision = 1 ORDER BY anio")]
    ws = wb.create_sheet("REVISION_MANUAL")
    _escribir_hoja(ws, cols_rev, filas)

    # ---- REGISTROS_SIN_DOCUMENTO ----
    cols_sin_doc = ["id_sistema", "anio", "codigo_original", "institucion", "tipo_instrumento", "objeto"]
    filas = [tuple(row[c] for c in cols_sin_doc) for row in conn.execute(
        f"SELECT {', '.join(cols_sin_doc)} FROM convenios WHERE estado_relacion_documental = 'NO_ENCONTRADA' ORDER BY anio")]
    ws = wb.create_sheet("REGISTROS_SIN_DOCUMENTO")
    _escribir_hoja(ws, cols_sin_doc, filas)

    # ---- DOCUMENTOS_SIN_REGISTRO ----
    cols_sin_reg = ["anio", "carpeta_tipo", "nombre", "extension", "tamano", "ruta"]
    filas = [tuple(row[c] for c in cols_sin_reg) for row in conn.execute(
        f"SELECT {', '.join(cols_sin_reg)} FROM documentos WHERE id_convenio IS NULL ORDER BY anio, carpeta_tipo")]
    ws = wb.create_sheet("DOCUMENTOS_SIN_REGISTRO")
    _escribir_hoja(ws, cols_sin_reg, filas)

    # ---- CATALOGO_TIPOS ----
    cols_cat = ["tipo_original", "tipo_normalizado", "clasificacion_general"]
    filas = [tuple(row[c] for c in cols_cat) for row in conn.execute(
        f"SELECT {', '.join(cols_cat)} FROM catalogo_tipos ORDER BY tipo_normalizado")]
    ws = wb.create_sheet("CATALOGO_TIPOS")
    _escribir_hoja(ws, cols_cat, filas)

    tablas_existentes = {r[0] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'")}

    # ---- EVIDENCIAS (Fase 3) ----
    if "evidencias_documentales" in tablas_existentes:
        cols_ev = ["id_evidencia", "id_convenio", "campo", "valor_extraido", "pagina", "fragmento_fuente",
                    "metodo_extraccion", "nivel_confianza", "fecha_analisis"]
        filas = [tuple(row[c] for c in cols_ev) for row in conn.execute(
            f"SELECT {', '.join(cols_ev)} FROM evidencias_documentales ORDER BY id_convenio")]
        ws = wb.create_sheet("EVIDENCIAS")
        _escribir_hoja(ws, cols_ev, filas)

    # ---- CONFLICTOS (Fase 3) ----
    if "conflicto_fecha" in _columnas_disponibles(conn, "convenios", ["conflicto_fecha"]):
        cols_conf = ["id_sistema", "anio", "codigo_original", "institucion", "fecha_suscripcion",
                      "fecha_suscripcion_documento", "fecha_finalizacion", "fecha_finalizacion_documento",
                      "estado_revision_vigencia", "ruta_documento_principal"]
        filas = [tuple(row[c] for c in cols_conf) for row in conn.execute(
            f"SELECT {', '.join(cols_conf)} FROM convenios WHERE conflicto_fecha='SI' ORDER BY anio")]
        ws = wb.create_sheet("CONFLICTOS")
        _escribir_hoja(ws, cols_conf, filas)

    conn.close()
    ruta_salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(ruta_salida))
